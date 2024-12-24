from robot.api.deco import keyword
from openpyxl import load_workbook

@keyword
def fetch_testdata_by_id(file_path, sheet_name, target_id):
    global workbook
    try:
        workbook = load_workbook(file_path)
        sheet = workbook[sheet_name]
        header = [cell.value for cell in sheet[1]]

        for row in sheet.iter_rows(min_row=2, values_only=True):
            if row[0] == target_id:
                data_dict = {}
                for col_name, value in zip(header, row):
                    if ',' in str(value):
                        data_dict[col_name] = [item.strip() for item in value.split(',')]
                    else:
                        data_dict[col_name] = value
                return data_dict
    except Exception as e:
        print(f"Error reading Excel file: {e}")
    finally:
        try:
            workbook.close()
        except NameError:
            pass
    return None


